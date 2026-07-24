from __future__ import annotations

from io import BytesIO
import re
from typing import Optional

import geopandas as gpd
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .analysis import DEFAULT_MATERIAL_STATE_PERCENT, build_condition_tables, condition_long_form, summarize_assignment
from .utils import APP_VERSION

AYA_BLUE = "002B5C"
AYA_GOLD = "C9A227"
LIGHT_BLUE = "EAF2F8"
LIGHT_GOLD = "FFF4CC"
LIGHT_GRAY = "F3F4F6"
RED = "D7191C"
YELLOW = "FDAE61"
GREEN = "1A9641"
WHITE = "FFFFFF"
BORDER = Side(style="thin", color="D9E2EC")


def _safe_sheet_name(name: object, used: set[str]) -> str:
    txt = str(name if name is not None else "Sin sistema").strip() or "Sin sistema"
    txt = re.sub(r"[\\/*?:\[\]]", "-", txt)
    txt = re.sub(r"\s+", " ", txt).strip()
    txt = txt[:31] or "Sistema"
    base = txt
    i = 2
    while txt in used:
        suffix = f"_{i}"
        txt = f"{base[:31-len(suffix)]}{suffix}"
        i += 1
    used.add(txt)
    return txt


def _write_dataframe(ws, df: pd.DataFrame, start_row: int, start_col: int, title: Optional[str] = None, percent_cols: Optional[set[str]] = None) -> int:
    """Write a dataframe with simple AyA styling. Returns next empty row."""
    row = start_row
    if title:
        ws.cell(row=row, column=start_col, value=title)
        title_cell = ws.cell(row=row, column=start_col)
        title_cell.font = Font(bold=True, color=AYA_BLUE, size=12)
        title_cell.fill = PatternFill("solid", fgColor=LIGHT_GOLD)
        row += 1

    if df is None or df.empty:
        ws.cell(row=row, column=start_col, value="Sin datos para mostrar.")
        return row + 2

    for j, col in enumerate(df.columns, start=start_col):
        cell = ws.cell(row=row, column=j, value=str(col))
        cell.fill = PatternFill("solid", fgColor=AYA_BLUE)
        cell.font = Font(bold=True, color=WHITE)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(top=BORDER, left=BORDER, right=BORDER, bottom=BORDER)
    row += 1

    percent_cols = percent_cols or set()
    for _, rec in df.iterrows():
        estado = str(rec.get("Estado", ""))
        for j, col in enumerate(df.columns, start=start_col):
            val = rec[col]
            cell = ws.cell(row=row, column=j)
            if pd.isna(val):
                cell.value = ""
            else:
                cell.value = val
            cell.border = Border(top=BORDER, left=BORDER, right=BORDER, bottom=BORDER)
            cell.alignment = Alignment(horizontal="right" if col != "Estado" else "left", vertical="center", wrap_text=True)
            if col == "Estado":
                cell.font = Font(bold=True)
                if estado == "Malo":
                    cell.fill = PatternFill("solid", fgColor="FDE2E2")
                elif estado == "Regular":
                    cell.fill = PatternFill("solid", fgColor="FFF1D6")
                elif estado == "Bueno":
                    cell.fill = PatternFill("solid", fgColor="DFF3E4")
                elif estado == "Total":
                    cell.fill = PatternFill("solid", fgColor=LIGHT_GRAY)
            else:
                if col in percent_cols or str(col).lower().startswith("porcentaje"):
                    cell.number_format = "0.0%"
                elif isinstance(val, (int, float)) and not isinstance(val, bool):
                    cell.number_format = '#,##0.0'
        row += 1
    return row + 2


def _apply_sheet_layout(ws, max_col: int) -> None:
    widths = {1: 18}
    for col_idx in range(2, max_col + 1):
        widths[col_idx] = 16
    for col_idx, width in widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.freeze_panes = "A4"
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = cell.alignment.copy(wrap_text=True, vertical="center")
    ws.sheet_view.showGridLines = False


def _write_parameters(ws, params: dict, start_row: int = 4) -> int:
    ws.cell(row=start_row, column=1, value="Parámetro").fill = PatternFill("solid", fgColor=AYA_BLUE)
    ws.cell(row=start_row, column=2, value="Valor").fill = PatternFill("solid", fgColor=AYA_BLUE)
    for c in (1, 2):
        ws.cell(row=start_row, column=c).font = Font(bold=True, color=WHITE)
        ws.cell(row=start_row, column=c).border = Border(top=BORDER, left=BORDER, right=BORDER, bottom=BORDER)
    row = start_row + 1
    for key, val in params.items():
        ws.cell(row=row, column=1, value=str(key))
        ws.cell(row=row, column=2, value=str(val))
        for c in (1, 2):
            ws.cell(row=row, column=c).border = Border(top=BORDER, left=BORDER, right=BORDER, bottom=BORDER)
            ws.cell(row=row, column=c).alignment = Alignment(wrap_text=True, vertical="top")
        row += 1
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 42
    return row + 2


def _assumptions_df() -> pd.DataFrame:
    df = pd.DataFrame(
        [
            {"Material": mat, "Malo": pct.get("Malo", 0.0), "Regular": pct.get("Regular", 0.0), "Bueno": pct.get("Bueno", 0.0)}
            for mat, pct in DEFAULT_MATERIAL_STATE_PERCENT.items()
        ]
    )
    return df


def _write_assumptions(ws, start_row: int, params: dict) -> int:
    title = ws.cell(row=start_row, column=1, value="Supuestos y consideraciones aplicadas")
    title.font = Font(bold=True, color=AYA_BLUE, size=12)
    title.fill = PatternFill("solid", fgColor=LIGHT_GOLD)
    start_row += 1
    items = [
        f"Las tuberías se segmentan con longitud máxima de {params.get('longitud_max_segmento_m', '')} m.",
        f"Las órdenes se asocian dentro de un radio de {params.get('radio_asociacion_m', '')} m.",
        "La asociación prioriza diámetro + sistema, luego diámetro y finalmente cercanía.",
        "Los materiales AC / asbesto-cemento, latón y otro se consideran por defecto 100 % en estado Malo.",
        "Cuando un segmento tiene órdenes asociadas, su estado se depura por el indicador de órdenes por cada 100 m.",
        "Cuando un segmento no tiene órdenes asociadas, se distribuye por los porcentajes base por material.",
        f"Umbrales activos: Bueno/Verde ≤ {params.get('umbral_verde_bueno', '')}; Regular/Amarillo hasta antes de {params.get('umbral_rojo_malo_desde', '')}; Malo/Rojo ≥ {params.get('umbral_rojo_malo_desde', '')} órdenes / 100 m.",
        "El resultado es una estimación de priorización y debe complementarse con criterio operativo, antigüedad, criticidad, inspección de campo, condición hidráulica y disponibilidad presupuestaria.",
    ]
    row = start_row
    for item in items:
        ws.cell(row=row, column=1, value="• " + item)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
        ws.cell(row=row, column=1).alignment = Alignment(wrap_text=True, vertical="top")
        row += 1
    row += 1
    df = _assumptions_df()
    return _write_dataframe(ws, df, row, 1, title="Porcentajes base por material", percent_cols={"Malo", "Regular", "Bueno"})


def _write_system_sheet(wb: Workbook, name: str, results: gpd.GeoDataFrame, system_col: Optional[str], material_col: Optional[str], params: dict) -> None:
    ws = wb.create_sheet(_safe_sheet_name(name, {s.title for s in wb.worksheets}))
    ws.cell(row=1, column=1, value=f"Estado estimado de tuberías - {name}")
    ws.cell(row=1, column=1).font = Font(bold=True, color=AYA_BLUE, size=16)
    ws.cell(row=2, column=1, value=f"Versión {APP_VERSION}")
    ws.cell(row=2, column=1).font = Font(color="666666")

    length_table, pct_table, _ = build_condition_tables(results, system_col=system_col, material_col=material_col, selected_system=name)
    row = 4
    row = _write_dataframe(ws, length_table, row, 1, title="Tabla 1. Longitud de tubería por estado y material", percent_cols={"Porcentaje"})
    row = _write_dataframe(ws, pct_table, row, 1, title="Tabla 2. Porcentaje de estado dentro de cada material", percent_cols=set([c for c in pct_table.columns if c != "Estado"]))
    row = _write_assumptions(ws, row, params)
    _apply_sheet_layout(ws, max_col=max(len(length_table.columns), len(pct_table.columns), 8))


def build_excel_report(
    results: gpd.GeoDataFrame,
    joined: Optional[gpd.GeoDataFrame],
    params: dict,
    system_col: Optional[str] = None,
    material_col: Optional[str] = None,
    selected_system: str = "Todos los sistemas",
) -> bytes:
    """Create an Excel report in memory with one worksheet per water-supply system."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Resumen"
    ws.sheet_view.showGridLines = False
    ws.cell(row=1, column=1, value="Reporte Excel - Estado estimado de tuberías")
    ws.cell(row=1, column=1).font = Font(bold=True, color=AYA_BLUE, size=16)
    ws.cell(row=2, column=1, value=f"Versión {APP_VERSION}")
    ws.cell(row=2, column=1).font = Font(color="666666")

    row = _write_parameters(ws, params, start_row=4)
    if joined is not None:
        summary = summarize_assignment(joined)
        if not summary.empty:
            row = _write_dataframe(ws, summary, row, 1, title="Resumen del método de asociación", percent_cols={"Porcentaje"})

    ws.cell(row=row, column=1, value="Nota sobre mapa")
    ws.cell(row=row, column=1).font = Font(bold=True, color=AYA_BLUE)
    ws.cell(row=row, column=1).fill = PatternFill("solid", fgColor=LIGHT_GOLD)
    row += 1
    ws.cell(row=row, column=1, value="El mapa interactivo se mantiene dentro del aplicativo Streamlit. El Excel concentra las tablas por sistema; no se inserta imagen del mapa para evitar dependencia de capturas externas.")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    ws.cell(row=row, column=1).alignment = Alignment(wrap_text=True, vertical="top")
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 42
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 18
    ws.column_dimensions["F"].width = 18
    ws.freeze_panes = "A4"

    # Determine which systems to export.
    long_df = condition_long_form(results, system_col=system_col, material_col=material_col)
    if selected_system and selected_system != "Todos los sistemas":
        systems = [selected_system]
    elif not long_df.empty:
        systems = sorted(long_df["sistema"].dropna().astype(str).unique().tolist())
    else:
        systems = ["Sin sistema"]

    for system in systems:
        _write_system_sheet(wb, system, results, system_col=system_col, material_col=material_col, params=params)

    # Final formatting on resumen.
    for row_cells in ws.iter_rows():
        for cell in row_cells:
            cell.alignment = cell.alignment.copy(wrap_text=True, vertical="center")
            if cell.row >= 4 and cell.value is not None:
                cell.border = Border(top=BORDER, left=BORDER, right=BORDER, bottom=BORDER)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()
